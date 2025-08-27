from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

IN_CSV  = Path("data/processed/it_companies_enriched.csv")
OUT_XLS = Path("outputs/IT_Companies_Leads.xlsx")
OUT_JSON = Path("outputs/IT_Companies_Leads.json")

def main():
    df = pd.read_csv(IN_CSV)

    # Reorder & rename for portfolio clarity
    cols = ["Company","Contact_URL","Phone","Email","Address","Website","Wikipedia"]
    df = df.reindex(columns=cols)

    # Simple quality: keep rows that have at least one of Phone/Email/Address
    base = len(df)
    df = df.dropna(subset=["Phone","Email","Address"], how="all")
    kept = len(df)

    # Sort: rows with Email first, then Phone
    df["has_email"] = df["Email"].notna()
    df["has_phone"] = df["Phone"].notna()
    df = df.sort_values(by=["has_email","has_phone","Company"], ascending=[False, False, True]).drop(columns=["has_email","has_phone"])

    # Save JSON (nice for portfolio)
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    df.to_json(OUT_JSON, orient="records", indent=2)

    # Save Excel
    OUT_XLS.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(OUT_XLS, index=False)

    # Light Excel formatting: bold header, freeze top row, set widths
    wb = load_workbook(OUT_XLS)
    ws = wb.active
    # Bold header
    for cell in ws[1]:
        cell.font = Font(bold=True)
    # Freeze header
    ws.freeze_panes = "A2"
    # Set column widths based on max content length
    for idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(idx)].width = min(max(12, max_len + 2), 60)
    wb.save(OUT_XLS)

    print(f"Excel saved to {OUT_XLS} | JSON saved to {OUT_JSON} | kept {kept}/{base} rows")

if __name__ == "__main__":
    main()
